import pandas as pd

#6 - py to exe

import os
import sys

application_path=os.path.dirname(sys.executable)

#1- Create a pivot table
df=pd.read_excel('supermarket_sales.xlsx')

df=df[['Gender','Product line','Total']]
pivot_table = df.pivot_table(index='Gender', columns='Product line', \
                             values='Total', aggfunc='sum').round(0)
pivot_table.to_excel('Output_Project_3_Create_a_pivot_table.xlsx', 'Report', startrow=4)



#2- Add a bar chart
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference 

wb = load_workbook('Output_Project_3_Create_a_pivot_table.xlsx')

sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print(min_column)
print(max_column)
print(min_row)
print(max_row)

barchart=BarChart()
#data (Right horizontal)
data = Reference(sheet, min_col=min_column+1,\
          max_col=max_column,\
          min_row=min_row,\
          max_row=max_row)

#categories (left vertical)- doesnt include the header 
categories = Reference(sheet, min_col=min_column,\
          max_col=min_column,\
          min_row=min_row+1,\
          max_row=max_row)


barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
sheet.add_chart(barchart, "B12")

barchart.title='sales by product line'
barchart.style=5
wb.save('Output_project_3_Add_a_bar_chart.xlsx')

#3 - write excel formulas with python
##sheet['B8']='=SUM(B6:B7)'
##sheet['B8'].style='Currency'
from openpyxl.utils import get_column_letter

for i in range(min_column+1, max_column+1):
    #print(i)
    letter = (get_column_letter(i))
    sheet[f'{letter}{max_row+1}']=f'=SUM({letter}{min_row+1}:{letter}{max_row})\n'
    sheet[f'{letter}{max_row+1}'].style='Currency'


wb.save('Output_project_3_write_excel_formula_with_python.xlsx')

#4 - Format cells

from openpyxl.styles import Font

month='January'
sheet['A1'] = 'Sales Report'
sheet['A2'] = month
sheet['A1'].font=Font('Arial', bold=True, size=20)
sheet['A2'].font=Font('Arial', bold=True, size=10)

wb.save(f'Output_project_3_Format_cells.xlsx')

#5- Create Pivot table into excel report
wb.save(f'Output_project_3_Create_Pivot_table_into_excel_report-{month}.xlsx')







